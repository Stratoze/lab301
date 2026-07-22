# -*- coding: utf-8 -*-
"""
Generate the populated HV_Test_Case.xlsx based on the SRS for "Do Ve So" project.
Run:  pip install openpyxl && python generate_testcase.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ================== STYLES ==================
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=HEADER_FILL):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        cell.alignment = WRAP


def auto_width(ws, cols, max_w=55):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max_w, 22)


# ================== COVER ==================
ws = wb.active
ws.title = "Cover"
cover = [
    ["TEST CASE"],
    [""],
    ["Project Name", "Do Ve So"],
    ["Project Code", "Lab301"],
    ["Document Code", "Lab301_TC_v1.0"],
    ["Issue Date", "22/07/2026"],
    ["Version", "1.0"],
    ["Creator", "Phan Dang Duy Phuc"],
    ["Reviewer/Approver", ""],
    [""],
    ["Record of change"],
    ["Effective Date", "Version", "Change Item", "Change description", "Reference"],
    ["22/07/2026", "1.0", "A", "Initial populated version based on SRS v1", "latest_HV_SRS.docx"],
]
for r, row in enumerate(cover, 1):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
ws["A1"].font = TITLE_FONT
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 25


# ================== TEST DESIGN ==================
ws = wb.create_sheet("TestDesign")
headers = ["Requirement Level 1", "Requirement Level 2", "Requirement Level 3",
           "Test Criteria", "Test Type", "Note"]
ws.append(headers)
style_header_row(ws, 1, len(headers))

design_rows = [
    # --- User Management ---
    ["User Management", "Registration", "Valid registration",
     "Register with valid full name, email, password, confirm password", "Function",
     "Default ROLE_USER, is_active=true, BCrypt password"],
    ["User Management", "Registration", "Required field validation",
     "Submit registration with missing required fields", "Function/GUI",
     "Red validation messages appear"],
    ["User Management", "Registration", "Invalid email format",
     "Enter invalid email (abc, abc@, abc@domain)", "Function",
     "Email format error"],
    ["User Management", "Registration", "Duplicate email",
     "Register with an email already in DB", "Function",
     "Duplicate email error"],
    ["User Management", "Registration", "Duplicate phone",
     "Register with phone already in DB (if phone unique)", "Function",
     "Duplicate phone error"],
    ["User Management", "Registration", "Password mismatch",
     "Password != Confirm Password", "Function/GUI",
     "Client-side validation error"],
    ["User Management", "Registration", "DB assertion after register",
     "Inspect users table after successful register", "Function/DB",
     "BCrypt hash, user_code=USR-MM-YYYY-NNNNNNNN, role=ROLE_USER"],
    ["User Management", "Login", "Valid login",
     "Login with correct email and password", "Function",
     "JWT returned, last_login updated, redirect to dashboard"],
    ["User Management", "Login", "Wrong password",
     "Login with correct email, wrong password", "Function",
     "Red error message, no JWT"],
    ["User Management", "Login", "Unknown email",
     "Login with non-existing email", "Function",
     "Error message, no JWT"],
    ["User Management", "Login", "Inactive account",
     "Login with blocked/inactive account", "Function",
     "Login denied, message account blocked"],
    ["User Management", "Login", "JWT expiry check",
     "Decode JWT after login", "Function/Security",
     "Expiry ~24h from issue"],
    ["User Management", "Forgot Password", "Valid recovery email",
     "Request reset for existing email", "Function",
     "Reset token created, email sent"],
    ["User Management", "Forgot Password", "Unknown email recovery",
     "Request reset for non-existing email", "Function",
     "Generic error or silent fail"],
    ["User Management", "Reset Password", "Valid reset",
     "Submit new password with valid token", "Function",
     "Password updated, token marked used, old password fails"],
    ["User Management", "Reset Password", "Expired/used token",
     "Reset with expired or already-used token", "Function/Security",
     "Reset denied"],
    ["User Management", "Reset Password", "Confirm mismatch",
     "New password != Confirm password", "Function/GUI",
     "Validation error"],
    ["User Management", "Change Password", "Logged-in change",
     "Change password with correct current password", "Function",
     "New password works, old fails"],
    ["User Management", "Change Password", "Wrong current password",
     "Change password with incorrect current password", "Function",
     "Error, password unchanged"],
    ["User Management", "Authorization", "USER access admin API",
     "Logged-in USER calls admin endpoint", "Security",
     "403 Forbidden"],
    ["User Management", "Authorization", "Unauthenticated API call",
     "Call protected API without JWT", "Security",
     "401 Unauthorized"],
    ["User Management", "Admin Search", "Keyword search",
     "Search users by name/email/phone/user_code", "Function",
     "List filtered correctly"],
    ["User Management", "Admin Pagination", ">20 users",
     "Open user list with >20 users", "Function/GUI",
     "Pagination, max 20 per page"],
    ["User Management", "Admin Edit", "Change role",
     "Edit user and change role USER<->ADMIN", "Function",
     "DB updated, permissions changed"],
    ["User Management", "Admin Edit", "Lock single user",
     "Lock an active user", "Function",
     "User blocked, cannot login"],
    ["User Management", "Admin Edit", "Unlock single user",
     "Unlock a blocked user", "Function",
     "User active again"],
    ["User Management", "Admin Bulk", "Bulk block",
     "Select multiple active users, bulk block", "Function",
     "All selected become blocked"],
    ["User Management", "Admin Bulk", "Bulk unlock",
     "Select multiple blocked users, bulk unlock", "Function",
     "All selected become active"],
    ["User Management", "Admin Optional", "Last login format (*)",
     "View last login column", "GUI",
     "Format dd/mm/yyyy"],
    ["User Management", "Admin Optional", "Filter inactive users (*)",
     "Filter users inactive 1w/1m/1q/1y", "Function",
     "Filter returns matching users"],
    ["User Management", "Admin Optional", "Send email (*)",
     "Select users, send custom email", "Function",
     "Email sent to selected users"],
    ["User Management", "Admin Optional", "Export Excel (*)",
     "Click Export on filtered list", "Function",
     "Excel downloaded with list"],
    ["User Management", "Admin Optional", "Google OAuth (*)",
     "Login with Google", "Function",
     "OAuth flow succeeds"],
    ["User Management", "Admin Optional", "Facebook OAuth (*)",
     "Login with Facebook", "Function",
     "OAuth flow succeeds"],
    # --- Ticket Management ---
    ["Ticket Management", "Create", "Valid creation",
     "Add ticket with station, date, prizes", "Function",
     "Saved as UNPUBLISH, result_code=RES-XXX-DDMMYYYY"],
    ["Ticket Management", "Create", "Missing required fields",
     "Submit without station/date/prizes", "Function/GUI",
     "Validation error, not saved"],
    ["Ticket Management", "Create", "Duplicate station+date",
     "Add result for same station and draw_date", "Function/DB",
     "Unique constraint violation"],
    ["Ticket Management", "Create", "Multi-line prizes",
     "Enter multiple winning numbers for one prize type", "Function",
     "Multiple prize_details rows created"],
    ["Ticket Management", "Create", "Non-numeric prize",
     "Enter letters/special chars in prize field", "Function/GUI",
     "Validation error: numbers only"],
    ["Ticket Management", "Create", "Duplicate prize number",
     "Same winning_number for same prize_type in same result", "Function/DB",
     "Unique constraint violation"],
    ["Ticket Management", "Edit", "Edit result",
     "Edit existing ticket and modify prizes", "Function",
     "Changes saved, updated_at updated"],
    ["Ticket Management", "Publish", "Publish result",
     "Toggle UNPUBLISH -> PUBLISH", "Function",
     "Status=PUBLISH, published_by/at saved, visible to users"],
    ["Ticket Management", "Publish", "Visibility check",
     "Guest/User selects published station/date", "Function",
     "Result available for checking"],
    ["Ticket Management", "Publish", "Unpublished not visible",
     "Guest/User selects UNPUBLISH station/date", "Function",
     "Not available for checking"],
    ["Ticket Management", "Search", "Filter by station+date",
     "Search results by station and date range", "Function",
     "Filtered list returned"],
    ["Ticket Management", "Search", "No match",
     "Search with criteria matching nothing", "GUI",
     "Empty state shown"],
    ["Ticket Management", "Pagination", ">10 results",
     "Open ticket list with >10 records", "Function/GUI",
     "Max 10 per page"],
    ["Ticket Management", "Audit", "total_queries increment",
     "After users check a published result, view count", "Function",
     "total_queries increases"],
    ["Ticket Management", "Authorization", "USER access admin ticket",
     "USER calls admin ticket API", "Security",
     "403 Forbidden"],
    ["Ticket Management", "Optional", "Preview ticket (*)",
     "Click Preview on a result", "GUI",
     "Popup with full prize details"],
    # --- Lottery Checking ---
    ["Lottery Checking", "Guest Check", "Winning ticket",
     "Guest checks winning ticket on published result", "Function",
     "Green result, prize shown, session.user_id=NULL"],
    ["Lottery Checking", "Guest Check", "Losing ticket",
     "Guest checks non-matching ticket", "Function",
     "Red result, total_won=0"],
    ["Lottery Checking", "Guest Check", "Guest multi-ticket rejected",
     "Guest tries to enter multiple tickets", "Function",
     "Only one ticket allowed or warning message"],
    ["Lottery Checking", "User Check", "Multi-ticket semicolon",
     "USER enters '123456;654321'", "Function",
     "Each ticket result listed"],
    ["Lottery Checking", "User Check", "Multi-ticket newline",
     "USER enters tickets separated by newline", "Function",
     "Each ticket result listed"],
    ["Lottery Checking", "User Check", "Special Prize win",
     "Ticket matches Special Prize", "Function",
     "Special Prize result, correct reward"],
    ["Lottery Checking", "User Check", "Lower prize win",
     "Ticket matches G8/G7", "Function",
     "Correct prize and reward"],
    ["Lottery Checking", "User Check", "Leading zeros",
     "Ticket with leading zeros '000123'", "Function",
     "Preserved and matched correctly"],
    ["Lottery Checking", "Validation", "Non-numeric ticket",
     "Enter 'ABC123' as ticket", "Function/GUI",
     "Validation error"],
    ["Lottery Checking", "Validation", "Unpublished result",
     "Try to check UNPUBLISH result", "Function",
     "Not allowed / no result message"],
    ["Lottery Checking", "Validation", "No result for station/date",
     "Select station/date with no result", "GUI",
     "Friendly 'no result' message"],
    ["Lottery Checking", "Performance", "50 tickets",
     "Check 50 tickets at once", "Non-function",
     "Response < 500ms"],
    ["Lottery Checking", "DB Assertion", "User session",
     "After USER checks, inspect DB", "Function/DB",
     "check_sessions row with user_id, check_histories rows per ticket"],
    ["Lottery Checking", "DB Assertion", "Guest session",
     "After guest checks, inspect DB", "Function/DB",
     "check_sessions row with user_id=NULL"],
    ["Lottery Checking", "Calculation", "total_spent",
     "Check total_spent calculation", "Function",
     "total_spent = n_tickets * 10000"],
    ["Lottery Checking", "Calculation", "total_won",
     "Check total_won calculation", "Function",
     "total_won = sum of winning amounts"],
    ["Lottery Checking", "History", "View history",
     "Logged-in user opens History page", "Function",
     "Previous checks listed with station/date/prize/amount/time"],
    ["Lottery Checking", "History", "Empty history",
     "Open history with no previous checks", "GUI",
     "Empty state shown"],
    ["Lottery Checking", "Analytics", "Bar chart",
     "Open Analytics page", "GUI/Function",
     "Bar chart shows total spent (red) vs total won (green)"],
    ["Lottery Checking", "Analytics", "Empty analytics",
     "Open analytics with no data", "GUI",
     "Empty chart/state, no crash"],
    ["Lottery Checking", "Responsive", "Mobile cards",
     "Open history/admin lists on mobile", "GUI",
     "Tables become card-based lists, readable"],
    ["Lottery Checking", "Responsive", "Mobile drawer",
     "Click hamburger icon on mobile", "GUI",
     "Side drawer opens with nav links"],
    ["Lottery Checking", "Optional", "Share to Facebook (*)",
     "Click Share after checking", "Function",
     "Facebook share dialog opens"],
    ["Lottery Checking", "Optional", "Gamification comments (*)",
     "Check near-miss result", "GUI",
     "Commentary shown based on result"],
    ["Lottery Checking", "Optional", "AdSense block (*)",
     "Guest views checking area", "GUI",
     "AdSense block shown"],
    ["Lottery Checking", "Optional", "History sort (*)",
     "Sort history asc/desc", "Function",
     "Order changes correctly"],
    # --- Security / Non-functional ---
    ["Security", "Password Storage", "BCrypt",
     "Inspect users.password after register", "Security",
     "BCrypt hash, not plaintext"],
    ["Security", "JWT", "Expiry",
     "Decode JWT", "Security",
     "Expiry ~24h"],
    ["Security", "Authorization", "401 without token",
     "Call protected API without token", "Security",
     "401"],
    ["Security", "Authorization", "403 USER on admin",
     "USER calls admin API", "Security",
     "403"],
    ["Security", "Authorization", "ADMIN inherits USER",
     "ADMIN uses user-level functions", "Function",
     "Allowed"],
    ["Security", "Pagination", "DB-level pagination",
     "Inspect admin list queries", "Function",
     "Pagination at DB level, not in-memory"],
    ["Security", "Error Handling", "No stack trace leak",
     "Trigger server error", "Security",
     "No stack trace / sensitive info in response"],
    ["Security", "Input Safety", "XSS",
     "Enter <script>alert(1)</script> in inputs", "Security",
     "Sanitized, no execution"],
    ["Security", "Input Safety", "SQL injection",
     "Enter SQL injection patterns in search/login", "Security",
     "JPA parameterized queries, injection fails"],
    ["Security", "CORS", "Origin check",
     "Inspect CORS/proxy config", "Security",
     "Only allowed origins"],
    # --- Code / UT / UI Checklist ---
    ["Code Quality", "Backend naming", "Java conventions",
     "Inspect Java classes/methods/vars", "Code",
     "UpperCamelCase classes, lowerCamelCase methods, UPPER_SNAKE_CASE constants"],
    ["Code Quality", "Frontend naming", "TS conventions",
     "Inspect React/TS files", "Code",
     "UpperCamelCase components, lowerCamelCase hooks, no any abuse"],
    ["Code Quality", "Structure", "Boilerplate",
     "Compare with SRS boilerplate", "Code",
     "Backend: config/controller/dto/entity/repository/security/service/utils. Frontend: api/components/pages/routes/utils"],
    ["Code Quality", "Dead code", "Unused removed",
     "Search unused classes/imports", "Code",
     "No dead code"],
    ["Code Quality", "Comments", "Complex logic",
     "Inspect checking algo, JWT filter, pagination", "Code",
     "Clear comments on complex logic"],
    ["Code Quality", "Build", "No errors",
     "Run backend + frontend build", "Code",
     "Both build without errors"],
    ["Unit Test", "DAO/Repository", "Repository tests",
     "Run repository tests", "Unit Test",
     "Tests exist for users, lottery_results, prize_details, check_sessions, check_histories"],
    ["Unit Test", "BO/Service", "Service tests",
     "Run service tests", "Unit Test",
     "Tests for auth, user mgmt, ticket mgmt, checking, history/analytics"],
    ["Unit Test", "Coverage", "JaCoCo report",
     "Run coverage tool", "Unit Test",
     "Coverage report generated, key logic covered"],
    ["System Test", "Test situations", "Documented",
     "Review test design", "System Test",
     "Positive/negative/auth/UI cases per function"],
    ["System Test", "Test cases", "Written",
     "Review test case list", "System Test",
     "All SRS functions covered"],
    ["UI", "Screens", "All screens",
     "Compare with SRS section 5", "GUI",
     "All 12 screens implemented"],
    ["UI", "Consistency", "Theme",
     "Inspect theme, spacing, radius", "GUI",
     "Light theme, AntD consistent, input 2px, button/card 12px radius"],
    ["UI", "Validation", "Form validation",
     "Test required/invalid/mismatch", "GUI",
     "Red validation messages consistent"],
    ["UI", "Responsive", "Mobile-first",
     "Test mobile + desktop", "GUI",
     "Mobile cards/drawer, desktop tables, no layout break"],
]

for r in design_rows:
    ws.append(r)
    style_data_row(ws, ws.max_row, len(headers))

auto_width(ws, len(headers))
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 55
ws.column_dimensions["F"].width = 50


# ================== FUNCTION SHEETS ==================
def build_function_sheet(name, module_code, requirement, cases):
    ws = wb.create_sheet(name)
    # Header block
    ws.append(["Module Code", module_code])
    ws.append(["Test requirement", requirement])
    ws.append(["Tester", "Phan Dang Duy Phuc"])
    ws.append(["Passed", 0, "Failed", 0, "Untested", len(cases),
               "Blocked", 0, "Skipped", 0, "Number of Test cases", len(cases)])
    ws.append([])  # blank row

    # Column headers
    cols = ["ID", "Test Case Description", "Test Case Procedure",
            "Expected Output", "Inter-test case Dependence",
            "Result (Pass/Fail/Untested/N/A)", "Test date", "Note"]
    ws.append(cols)
    style_header_row(ws, 6, len(cols))

    for case in cases:
        ws.append(case)
        style_data_row(ws, ws.max_row, len(cols))
        # Highlight Untested
        ws.cell(row=ws.max_row, column=6).fill = YELLOW_FILL

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 35
    return ws


# ---------- Function 1: User Management ----------
f1_cases = [
    ["F1-01", "Register with valid data",
     "1. Open Register tab. 2. Enter Full Name, Valid Email, Password, Confirm Password. 3. Click Sign Up.",
     "User created. Redirect to Login. DB: role=ROLE_USER, is_active=true, password BCrypt, user_code=USR-MM-YYYY-NNNNNNNN.",
     "None", "Untested", "", "Positive"],
    ["F1-02", "Register missing required fields",
     "1. Leave Full Name/Email/Password blank. 2. Click Sign Up.",
     "Red validation messages appear for each required field.",
     "None", "Untested", "", "Negative"],
    ["F1-03", "Register invalid email format",
     "1. Enter 'abc', 'abc@', 'abc@domain' as email. 2. Click Sign Up.",
     "Email format validation error appears.",
     "None", "Untested", "", "Negative"],
    ["F1-04", "Register duplicate email",
     "1. Use email already registered. 2. Click Sign Up.",
     "Duplicate email error. Account not created.",
     "Existing user", "Untested", "", "Negative"],
    ["F1-05", "Register duplicate phone",
     "1. Use phone already registered. 2. Click Sign Up.",
     "Duplicate phone error (if phone unique).",
     "Existing user with phone", "Untested", "", "Negative"],
    ["F1-06", "Register password mismatch",
     "1. Password='123456'. 2. Confirm='654321'. 3. Click Sign Up.",
     "Error: passwords do not match.",
     "None", "Untested", "", "Negative"],
    ["F1-07", "Login success",
     "1. Enter valid Email and Password. 2. Click Sign In.",
     "JWT generated and stored in LocalStorage. Redirect to Dashboard. last_login updated in DB.",
     "[F1-01]", "Untested", "", "Positive"],
    ["F1-08", "Login wrong password",
     "1. Enter valid Email, wrong Password. 2. Click Sign In.",
     "Red error message. No JWT stored.",
     "None", "Untested", "", "Negative"],
    ["F1-09", "Login unknown email",
     "1. Enter non-existing email. 2. Click Sign In.",
     "Error message. No JWT stored.",
     "None", "Untested", "", "Negative"],
    ["F1-10", "Login inactive account",
     "1. Admin blocks user. 2. User tries login.",
     "Login denied. Message: account blocked.",
     "Admin blocked user", "Untested", "", "Negative"],
    ["F1-11", "JWT expiry check",
     "1. Login. 2. Decode JWT.",
     "JWT expiry ~24h from issue.",
     "[F1-07]", "Untested", "", "Security"],
    ["F1-12", "Forgot password valid email",
     "1. Open Reset Modal. 2. Enter existing email. 3. Click Send.",
     "Success message. Reset token created in DB. Email sent.",
     "None", "Untested", "", "Positive"],
    ["F1-13", "Forgot password unknown email",
     "1. Enter non-existing email. 2. Click Send.",
     "Generic error or silent fail. No usable token.",
     "None", "Untested", "", "Negative"],
    ["F1-14", "Reset password valid token",
     "1. Use valid reset token. 2. Enter new password + confirm. 3. Submit.",
     "Password updated. Token marked used. Old password fails, new works.",
     "[F1-12]", "Untested", "", "Positive"],
    ["F1-15", "Reset password expired/used token",
     "1. Use expired or already-used token. 2. Submit new password.",
     "Reset denied. Error message.",
     "Expired/used token", "Untested", "", "Security"],
    ["F1-16", "Reset password confirm mismatch",
     "1. New password != Confirm password. 2. Submit.",
     "Validation error. Password not updated.",
     "[F1-12]", "Untested", "", "Negative"],
    ["F1-17", "Change password logged-in",
     "1. Logged-in user. 2. Enter correct current password + new password. 3. Submit.",
     "Password changed. New works, old fails.",
     "[F1-07]", "Untested", "", "Positive"],
    ["F1-18", "Change password wrong current",
     "1. Enter wrong current password. 2. Submit.",
     "Error. Password unchanged.",
     "[F1-07]", "Untested", "", "Negative"],
    ["F1-19", "Protected API without token",
     "1. Call GET /users/me without JWT.",
     "401 Unauthorized.",
     "None", "Untested", "", "Security"],
    ["F1-20", "Protected API with invalid JWT",
     "1. Call protected API with invalid/expired JWT.",
     "401 Unauthorized. Redirect to login.",
     "Invalid token", "Untested", "", "Security"],
    ["F1-21", "USER access admin endpoint",
     "1. Login as USER. 2. Call admin API.",
     "403 Forbidden.",
     "[F1-07]", "Untested", "", "Security"],
    ["F1-22", "Admin search user by keyword",
     "1. Login as ADMIN. 2. Go to Manage User. 3. Type 'Nguyen' in search. 4. Enter.",
     "List filtered to users containing 'Nguyen'.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-23", "Admin pagination >20 users",
     "1. Have >20 users in DB. 2. Open user list.",
     "Max 20 users per page. Pagination controls work.",
     "Admin login, >20 users", "Untested", "", "Positive"],
    ["F1-24", "Admin change user role",
     "1. Edit user. 2. Change role USER<->ADMIN. 3. Save.",
     "Role updated in DB. Permissions changed.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-25", "Admin lock single user",
     "1. Select active user. 2. Click Lock.",
     "User blocked. Cannot login.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-26", "Admin unlock single user",
     "1. Select blocked user. 2. Click Unlock.",
     "User active. Can login.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-27", "Admin bulk block",
     "1. Select multiple active users. 2. Bulk Block.",
     "All selected become blocked. Tags turn red.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-28", "Admin bulk unlock",
     "1. Select multiple blocked users. 2. Bulk Unlock.",
     "All selected become active.",
     "Admin login", "Untested", "", "Positive"],
    ["F1-29", "Admin last login format (*)",
     "1. View user list.",
     "Last login shown as dd/mm/yyyy.",
     "Admin login", "Untested", "", "Optional"],
    ["F1-30", "Admin filter inactive users (*)",
     "1. Filter users inactive 1w/1m/1q/1y.",
     "Filter returns matching users.",
     "Admin login", "Untested", "", "Optional"],
    ["F1-31", "Admin send email (*)",
     "1. Select users. 2. Send Email. 3. Enter subject+content. 4. Send.",
     "Email sent to selected users. Confirmation shown.",
     "Admin login, mail configured", "Untested", "", "Optional"],
    ["F1-32", "Admin export Excel (*)",
     "1. Click Export on filtered list.",
     "Excel file downloaded with displayed users.",
     "Admin login", "Untested", "", "Optional"],
    ["F1-33", "Google OAuth login (*)",
     "1. Click Login with Google.",
     "OAuth flow succeeds. User logged in.",
     "Google OAuth configured", "Untested", "", "Optional"],
    ["F1-34", "Facebook OAuth login (*)",
     "1. Click Login with Facebook.",
     "OAuth flow succeeds. User logged in.",
     "Facebook OAuth configured", "Untested", "", "Optional"],
]

build_function_sheet("Function_1", "Module1",
                     "Testing user lifecycle from registration to administrative management.",
                     f1_cases)


# ---------- Function 2: Ticket Management ----------
f2_cases = [
    ["F2-01", "Add ticket valid",
     "1. Admin login. 2. Click Add Ticket. 3. Select station, draw date. 4. Enter prize numbers. 5. Save.",
     "Result saved as UNPUBLISH. result_code=RES-XXX-DDMMYYYY. created_by/at saved.",
     "Admin login", "Untested", "", "Positive"],
    ["F2-02", "Add ticket missing required fields",
     "1. Leave station/date/prizes blank. 2. Save.",
     "Validation error. Not saved.",
     "Admin login", "Untested", "", "Negative"],
    ["F2-03", "Add ticket duplicate station+date",
     "1. Add result for existing station+draw_date.",
     "Unique constraint violation. Error message.",
     "Existing result", "Untested", "", "Negative"],
    ["F2-04", "Add ticket multi-line prizes",
     "1. Enter multiple lines for 4th Prize. 2. Save.",
     "Multiple prize_details rows created for that prize type.",
     "Admin login", "Untested", "", "Positive"],
    ["F2-05", "Add ticket non-numeric prize",
     "1. Enter 'ABC' or '12A' in prize field. 2. Save.",
     "Validation error: numbers only.",
     "Admin login", "Untested", "", "Negative"],
    ["F2-06", "Add ticket duplicate prize number",
     "1. Enter same winning_number for same prize_type in same result.",
     "Unique constraint violation.",
     "Admin login", "Untested", "", "Negative"],
    ["F2-07", "Edit ticket",
     "1. Edit existing ticket. 2. Modify prizes. 3. Save.",
     "Changes saved. updated_at updated.",
     "Existing result", "Untested", "", "Positive"],
    ["F2-08", "Publish result",
     "1. Edit UNPUBLISH ticket. 2. Toggle to PUBLISH. 3. Save.",
     "Status=PUBLISH. published_by/at saved. Visible to users.",
     "Admin login", "Untested", "", "Positive"],
    ["F2-09", "Published result visible to users",
     "1. User selects published station/date.",
     "Result available for checking.",
     "[F2-08]", "Untested", "", "Positive"],
    ["F2-10", "Unpublished result not visible",
     "1. User selects UNPUBLISH station/date.",
     "Not available for checking.",
     "UNPUBLISH exists", "Untested", "", "Negative"],
    ["F2-11", "Search ticket by station+date",
     "1. Filter by station and date range.",
     "Filtered list returned.",
     "Admin login", "Untested", "", "Positive"],
    ["F2-12", "Search ticket no match",
     "1. Filter with criteria matching nothing.",
     "Empty state (Icon + 'No Data').",
     "Admin login", "Untested", "", "Negative"],
    ["F2-13", "Ticket pagination >10",
     "1. Have >10 results. 2. Open ticket list.",
     "Max 10 per page. Pagination works.",
     "Admin login, >10 results", "Untested", "", "Positive"],
    ["F2-14", "total_queries increment",
     "1. Users check a published result. 2. Admin views result.",
     "total_queries increases.",
     "Published result checked", "Untested", "", "Positive"],
    ["F2-15", "USER access admin ticket API",
     "1. Login as USER. 2. Call admin ticket API.",
     "403 Forbidden.",
     "[F1-07]", "Untested", "", "Security"],
    ["F2-16", "Preview ticket (*)",
     "1. Click Preview on a result.",
     "Popup with full prize details.",
     "Admin login", "Untested", "", "Optional"],
]

build_function_sheet("Function_2", "Module2",
                     "Testing lottery result entry, numeric validation, and publishing.",
                     f2_cases)


# ---------- Function 3: Lottery Checking ----------
f3_cases = [
    ["F3-01", "Guest winning ticket",
     "1. As guest, select published station/date. 2. Enter winning ticket. 3. Check.",
     "Green result. Prize type and reward shown. session.user_id=NULL. total_queries++.",
     "Published result", "Untested", "", "Positive"],
    ["F3-02", "Guest losing ticket",
     "1. As guest, enter non-matching ticket. 2. Check.",
     "Red result: 'Better luck next time'. total_won=0. Session saved.",
     "Published result", "Untested", "", "Positive"],
    ["F3-03", "Guest multi-ticket rejected",
     "1. As guest, enter multiple tickets separated by ; or newline.",
     "Only one ticket allowed, or warning message.",
     "Published result", "Untested", "", "Negative"],
    ["F3-04", "User multi-ticket semicolon",
     "1. Login as USER. 2. Enter '123456;654321'. 3. Check.",
     "Each ticket result listed.",
     "User login", "Untested", "", "Positive"],
    ["F3-05", "User multi-ticket newline",
     "1. Login as USER. 2. Enter tickets separated by newline. 3. Check.",
     "Each ticket result listed.",
     "User login", "Untested", "", "Positive"],
    ["F3-06", "Special Prize win",
     "1. Enter ticket matching Special Prize.",
     "Green alert: 'Congratulations'. Special Prize + correct reward.",
     "Published result with Special Prize", "Untested", "", "Positive"],
    ["F3-07", "Lower prize win (G8/G7)",
     "1. Enter ticket matching G8/G7.",
     "Correct lower prize and reward shown.",
     "Published result with lower prize", "Untested", "", "Positive"],
    ["F3-08", "Leading zeros ticket",
     "1. Enter '000123'. 2. Check.",
     "Leading zeros preserved and matched correctly.",
     "Published result with leading-zero prize", "Untested", "", "Positive"],
    ["F3-09", "Non-numeric ticket",
     "1. Enter 'ABC123'. 2. Check.",
     "Validation error. Not executed.",
     "None", "Untested", "", "Negative"],
    ["F3-10", "Check UNPUBLISH result",
     "1. Try to check UNPUBLISH result.",
     "Not allowed / no result message.",
     "UNPUBLISH result", "Untested", "", "Negative"],
    ["F3-11", "No result for station/date",
     "1. Select station/date with no result.",
     "Friendly 'no result' message.",
     "No result", "Untested", "", "Negative"],
    ["F3-12", "50 tickets performance",
     "1. Enter 50 valid tickets. 2. Check.",
     "Response < 500ms. All results correct.",
     "Published result", "Untested", "", "Non-functional"],
    ["F3-13", "DB assertion user session",
     "1. USER checks tickets. 2. Inspect DB.",
     "One check_sessions row with user_id. Multiple check_histories rows per ticket.",
     "User login", "Untested", "", "DB"],
    ["F3-14", "DB assertion guest session",
     "1. Guest checks ticket. 2. Inspect DB.",
     "check_sessions row with user_id=NULL.",
     "None", "Untested", "", "DB"],
    ["F3-15", "total_spent calculation",
     "1. Check session created. 2. Inspect total_spent.",
     "total_spent = n_tickets * 10000.",
     "Check session", "Untested", "", "Positive"],
    ["F3-16", "total_won calculation",
     "1. At least one winning ticket. 2. Inspect total_won.",
     "total_won = sum of winning amounts.",
     "Winning session", "Untested", "", "Positive"],
    ["F3-17", "View history",
     "1. USER opens History page.",
     "Previous checks listed: station/date/ticket/prize/amount/time.",
     "User login, previous checks", "Untested", "", "Positive"],
    ["F3-18", "Empty history",
     "1. USER with no history opens History page.",
     "Empty state shown.",
     "User login, no history", "Untested", "", "Negative"],
    ["F3-19", "Analytics bar chart",
     "1. USER opens Analytics page.",
     "Bar chart shows total spent (red) vs total won (green). Values match check_sessions.",
     "User login, sessions exist", "Untested", "", "GUI"],
    ["F3-20", "Empty analytics",
     "1. USER with no data opens Analytics.",
     "Empty chart/state, no crash.",
     "User login, no sessions", "Untested", "", "GUI"],
    ["F3-21", "Mobile card view",
     "1. Open history/admin lists on mobile viewport.",
     "Tables become card-based lists, readable.",
     "Mobile viewport", "Untested", "", "GUI"],
    ["F3-22", "Mobile navigation drawer",
     "1. Click hamburger icon on mobile top bar.",
     "Side drawer opens with navigation links.",
     "Mobile viewport", "Untested", "", "GUI"],
    ["F3-23", "Share to Facebook (*)",
     "1. Click Share after checking.",
     "Facebook share dialog opens.",
     "Facebook SDK configured", "Untested", "", "Optional"],
    ["F3-24", "Gamification comments (*)",
     "1. Check near-miss result.",
     "Commentary shown based on result.",
     "Published result", "Untested", "", "Optional"],
    ["F3-25", "AdSense block (*)",
     "1. Guest views checking area.",
     "AdSense block shown.",
     "AdSense configured", "Untested", "", "Optional"],
    ["F3-26", "History sort (*)",
     "1. Sort history asc/desc.",
     "Order changes correctly.",
     "History exists", "Untested", "", "Optional"],
]

build_function_sheet("Function_3", "Module3",
                     "Testing lottery checking algorithm, visualization, and responsive design.",
                     f3_cases)


# ================== TEST CASE LIST (Index) ==================
ws = wb.create_sheet("Test case List")
ws.insert_rows(1, 5)
ws["A1"].value = "TEST CASE LIST"
ws["A1"].font = TITLE_FONT
ws["A3"].value = "Project Name"
ws["B3"].value = "Do Ve So"
ws["A4"].value = "Project Code"
ws["B4"].value = "Lab301"
ws["A5"].value = "Test Environment"
ws["B5"].value = "1. Server: Tomcat (Spring Boot, Java 21)\n2. Database: MySQL 8.0\n3. Browser: Chrome, Safari (Desktop & Mobile)"
ws["B5"].alignment = WRAP

ws.append([])
ws.append(["No", "Function Name", "Sheet Name", "Description", "Pre-Condition"])
style_header_row(ws, ws.max_row, 5)

index_rows = [
    [1, "Function 1: User Management", "Function_1",
     "Testing Registration, Login, Forgot/Reset/Change Password, Authorization, Admin User controls.",
     "System deployed. DB connected."],
    [2, "Function 2: Ticket Management", "Function_2",
     "Testing admin creation, editing, publishing, searching, pagination of lottery results.",
     "Logged in as ADMIN."],
    [3, "Function 3: Lottery Checking", "Function_3",
     "Testing checking algorithm (guest/user), history, analytics, responsive UI, DB assertions.",
     "Published lottery results exist in DB."],
]
for r in index_rows:
    ws.append(r)
    style_data_row(ws, ws.max_row, 5)

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 40


# ================== TEST REPORT ==================
ws = wb.create_sheet("Test Report")
ws["A1"].value = "TEST REPORT"
ws["A1"].font = TITLE_FONT
ws.append(["Project Name", "Do Ve So"])
ws.append(["Project Code", "Lab301"])
ws.append(["Document Code", "Lab301_TestReport_v1.0"])
ws.append(["Issue Date", "22/07/2026"])
ws.append(["Notes", "Release 1 includes 3 modules: Module1, Module2, Module3"])
ws.append([])

headers = ["No", "Module code", "Pass", "Fail", "Untested", "Blocked", "Skipped", "Number of test cases"]
ws.append(headers)
style_header_row(ws, ws.max_row, len(headers))

f1_count = len(f1_cases)
f2_count = len(f2_cases)
f3_count = len(f3_cases)
total = f1_count + f2_count + f3_count

report_rows = [
    [1, "Module1", 0, 0, f1_count, 0, 0, f1_count],
    [2, "Module2", 0, 0, f2_count, 0, 0, f2_count],
    [3, "Module3", 0, 0, f3_count, 0, 0, f3_count],
    ["", "Sub total", 0, 0, total, 0, 0, total],
]
for r in report_rows:
    ws.append(r)
    style_data_row(ws, ws.max_row, len(headers))

ws.append([])
ws.append(["", "Test coverage", "", f"{total/total*100:.2f}", "%", "", "", ""])
ws.append(["", "Test successful coverage", "", "0.00", "%", "", "", "(until actual execution)"])

for c in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(c)].width = 20


# ================== SAVE ==================
out = "HV_Test_Case_Populated.xlsx"
wb.save(out)
print(f"[OK] Generated: {out}")
print(f"     Function_1 cases: {f1_count}")
print(f"     Function_2 cases: {f2_count}")
print(f"     Function_3 cases: {f3_count}")
print(f"     TOTAL cases     : {total}")
